import streamlit as st
import pdfplumber
import pandas as pd
import re
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Extracteur Pro 1xBet", layout="wide", page_icon="⚽")

st.markdown("""
    <style>
    .main-title { font-size: 2.2rem; font-weight: bold; color: #1E3A8A; margin-bottom: 0.5rem; }
    .sub-title { font-size: 1.1rem; color: #4B5563; margin-bottom: 2rem; }
    </style>
""", unsafe_style_allowed=True)

st.markdown('<div class="main-title">⚽ Extracteur de Matchs 1xBet Pro</div>', unsafe_style_allowed=True)
st.markdown('<div class="sub-title">Algorithme de parsing intelligent avec nettoyage automatique des cotes et mise en forme Excel.</div>', unsafe_style_allowed=True)

uploaded_file = st.file_uploader("Glissez votre fichier PDF exporté de 1xBet ici", type="pdf")

def nettoyer_nom_equipe(nom):
    """Supprime les cotes et les en-têtes de marchés (1, X, 2, 1X...) collés au nom de l'équipe"""
    # Supprime les cotes décimales (ex: 2.3, 3.75, etc.) et tout ce qui suit
    nom = re.sub(r'\b\d+[\.,]\d+\b.*', '', nom)
    # Supprime les indicateurs de colonnes de paris isolés
    nom = re.sub(r'\b(1|X|2|1X|12|2X|\+\d+)\b', '', nom)
    # Nettoie les caractères de ponctuation résiduels en fin de chaîne
    nom = re.sub(r'[\s\-\+\,\.\|\/]+$', '', nom)
    return nom.strip()

def extraire_donnees_1xbet(pdf_file):
    matchs = []
    current_competition = "Général / Autre"
    
    # Mots-clés indiquant une ligne de compétition
    comp_keywords = ["Championnat", "Coupe", "Ligue", "Copa", "Matchs amicaux", "EMF Mini", "Division", "Clubs", "Qualif", "Liga"]
    # Mots-clés pour ignorer le bruit (boutons, textes de l'interface)
    junk_keywords = ["INSCRIPTION", "SE CONNECTER", "BONUS", "COUPON", "TOTAL", "PLUS DE", "HANDICAP", "PARI COMBINÉ"]

    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
                
            # Découpage en lignes propres
            lines = [line.strip() for line in text.split("\n") if line.strip()]
            
            i = 0
            while i < len(lines):
                line_str = lines[i]
                
                # Détection de la ligne de Date / Heure (ex: 27/05 09:30 ou 27/05/10:00)
                date_match = re.search(r'(\d{2}/\d{2})[\s/:]+(\d{2}:\d{2})', line_str)
                
                if date_match:
                    # RECHERCHE INVERSÉE : Les équipes se trouvent AU-DESSUS de la date
                    teams_found = []
                    j = i - 1
                    
                    while j >= 0 and len(teams_found) < 2:
                        prev_line = lines[j]
                        
                        is_prev_comp = any(k in prev_line for k in comp_keywords)
                        is_prev_junk = any(jk in prev_line.upper() for jk in junk_keywords)
                        is_prev_date = re.search(r'\d{2}/\d{2}', prev_line)
                        
                        # Si la ligne du dessus n'est ni une compète, ni du junk, ni une autre date, c'est une équipe !
                        if not is_prev_comp and not is_prev_junk and not is_prev_date and len(prev_line) > 2:
                            teams_found.append(prev_line)
                        j -= 1
                    
                    # Si on a trouvé au moins 2 lignes d'équipes valides au-dessus
                    if len(teams_found) >= 2:
                        # Comme on a scanné vers le haut : teams_found[0] = Extérieur, teams_found[1] = Domicile
                        equipe_ext_raw = teams_found[0]
                        equipe_dom_raw = teams_found[1]
                        
                        # Collecte du bloc de texte pour extraire les cotes proprement (1, X, 2)
                        context_lines = [equipe_dom_raw, equipe_ext_raw, line_str]
                        if i + 1 < len(lines): context_lines.append(lines[i+1])
                        if i + 2 < len(lines): context_lines.append(lines[i+2])
                        
                        combined_context = " ".join(context_lines)
                        all_floats = re.findall(r'\b\d+[\.,]\d+\b', combined_context)
                        all_floats = [f.replace(',', '.') for f in all_floats]
                        
                        # Nettoyage final des noms
                        equipe_dom = nettoyer_nom_equipe(equipe_dom_raw)
                        equipe_ext = nettoyer_nom_equipe(equipe_ext_raw)
                        
                        # Formatage des cotes
                        cotes_str = " / ".join(all_floats[:3]) if len(all_floats) >= 3 else "N/A"
                        
                        # Sécurité pour éviter d'ajouter des lignes vides ou mal extraites
                        if equipe_dom and equipe_ext and equipe_dom != equipe_ext:
                            matchs.append({
                                "🏠 Équipe domicile": equipe_dom,
                                "✈️ Équipe extérieur": equipe_ext,
                                "🏆 Compétition": current_competition,
                                "🕒 Coup d'envoi": f"{date_match.group(1)} {date_match.group(2)}",
                                "💰 Cotes (1 / X / 2)": cotes_str
                            })
                else:
                    # Mise à jour de la compétition courante si on croise une ligne d'en-tête
                    if any(k in line_str for k in comp_keywords):
                        comp_clean = re.sub(r'\b(1|X|2|1X|12|2X|\+\d+)\b.*', '', line_str)
                        comp_clean = re.sub(r'[\s\-\+\,\.\|\/]+$', '', comp_clean).strip()
                        if len(comp_clean) > 5:
                            current_competition = comp_clean
                
                i += 1
                
    return matchs

if uploaded_file is not None:
    with st.spinner("Analyse et réalignement des données en cours..."):
        donnees = extraire_donnees_1xbet(uploaded_file)
        
        if donnees:
            df = pd.DataFrame(donnees)
            
            # Suppression des doublons potentiels de parsing
            df.drop_duplicates(subset=["🏠 Équipe domicile", "✈️ Équipe extérieur", "🕒 Coup d'envoi"], inplace=True)
            
            st.success(f"🔥 {len(df)} matchs réalignés et extraits avec succès !")
            
            # Aperçu sur Streamlit
            st.dataframe(df, use_container_width=True)
            
            # Génération d'un fichier Excel stylisé professionnellement
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Data_Matches')
                
                # Design de la feuille Excel
                workbook = writer.book
                worksheet = writer.sheets['Data_Matches']
                
                # Styles
                header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
                header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                data_font = Font(name="Arial", size=10)
                zebra_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                thin_border = Border(
                    left=Side(style='thin', color='E5E7EB'),
                    right=Side(style='thin', color='E5E7EB'),
                    top=Side(style='thin', color='E5E7EB'),
                    bottom=Side(style='thin', color='E5E7EB')
                )
                
                # Appliquer le style aux en-têtes
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Appliquer le style aux lignes de données
                for row_num in range(2, len(df) + 2):
                    is_even = (row_num % 2 == 0)
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.font = data_font
                        cell.border = thin_border
                        
                        # Aligner selon la colonne
                        if col_num in [4, 5]: # Date et Cotes
                            cell.alignment = Alignment(horizontal="center")
                        else:
                            cell.alignment = Alignment(horizontal="left")
                            
                        # Zebra striping
                        if is_even:
                            cell.fill = zebra_fill
                
                # Ajustement automatique de la largeur des colonnes
                for col in worksheet.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
            st.download_button(
                label="📥 Télécharger le fichier Excel corrigé",
                data=buffer.getvalue(),
                file_name="extraction_1xbet_corrigee.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("Aucun match trouvé. Assurez-vous que le PDF provient bien d'un export de ligne 1xBet.")
